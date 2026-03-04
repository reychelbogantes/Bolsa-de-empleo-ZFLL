from django.contrib.auth import get_user_model
from django.db import transaction
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView

from shared.permissions import IsInstitucion
from openpyxl import load_workbook

from apps.accounts.models import PerfilPracticante
from apps.institutions.models import ProgramaFormacion
from .models import LoteImportacion, EstudianteImportado
from .serializers import LoteImportacionSerializer

User = get_user_model()

DEFAULT_PASSWORD = "estudiante123"

LABEL_TO_KEY = {
    "Email": "correo",
    "Nombre": "nombre",
    "Apellido": "apellido",
    "Username": "username",
    "Teléfono": "telefono",
    "ID Institución": "institucion_id",
    "ID Programa": "programa_id",
    "Fecha Inicio (YYYY-MM-DD)": "periodo_inicio",
    "Fecha Fin (YYYY-MM-DD)": "periodo_fin",
    "Horas Requeridas": "horas_requeridas",
}


def _unique_username(base: str) -> str:
    base = (base or "user").strip().replace(".", "_").replace("-", "_")
    base = base[:20] if base else "user"
    username = base
    if not User.objects.filter(username=username).exists():
        return username
    i = 2
    while User.objects.filter(username=f"{base}_{i}").exists():
        i += 1
    return f"{base}_{i}"


def _cell(v):
    return "" if v is None else str(v).strip()


@transaction.atomic
def procesar_lote_excel(lote: LoteImportacion):
    """
    Procesa el Excel del lote:
    - Crea/actualiza usuarios (email único)
    - Asigna rol 'practicante'
    - Crea/actualiza PerfilPracticante
    - Guarda resultados por fila en lote.log_errores (también incluye éxitos)
    """
    lote.estado = "processing"
    lote.log_errores = []
    lote.total_registros = 0
    lote.creados = 0
    lote.actualizados = 0
    lote.con_error = 0
    lote.save(update_fields=["estado", "log_errores", "total_registros", "creados", "actualizados", "con_error"])

    # abrir archivo desde FileField
    with lote.archivo.open("rb") as f:
        wb = load_workbook(f, data_only=True)
    ws = wb.active

    # headers: fila 1
    headers = [(_cell(c.value)) for c in ws[1]]
    headers = [h for h in headers if h]  # recorta vacíos al final

    # datos desde fila 3 (fila 2 es ejemplo)
    start_row = 3

    resultados = []
    total = 0

    for r in range(start_row, ws.max_row + 1):
        row_vals = [_cell(ws.cell(row=r, column=c + 1).value) for c in range(len(headers))]
        if not any(v for v in row_vals):
            continue  # fila vacía

        total += 1
        obj = {"fila": r, "estado": None, "msg": "", "email": "", "username": ""}

        # construir registro
        reg = {}
        for i, h in enumerate(headers):
            key = LABEL_TO_KEY.get(h)
            if key:
                reg[key] = row_vals[i] if i < len(row_vals) else ""

        email = (reg.get("correo") or "").lower().strip()
        nombre = (reg.get("nombre") or "").strip()
        apellido = (reg.get("apellido") or "").strip()
        telefono = (reg.get("telefono") or "").strip()

        obj["email"] = email

        # validaciones mínimas
        errores = []
        if not email:
            errores.append("Email requerido")
        if not nombre:
            errores.append("Nombre requerido")

        if errores:
            lote.con_error += 1
            obj["estado"] = "error"
            obj["msg"] = "; ".join(errores)
            resultados.append(obj)
            continue

        # username
        username_in = (reg.get("username") or "").strip()
        base_un = username_in or email.split("@")[0]
        username = _unique_username(base_un)
        obj["username"] = username

        # programa (opcional)
        programa_id = (reg.get("programa_id") or "").strip()
        programa = None
        if programa_id:
            try:
                programa = ProgramaFormacion.objects.get(id=int(programa_id), institucion=lote.institucion)
            except Exception:
                lote.con_error += 1
                obj["estado"] = "error"
                obj["msg"] = f"ID Programa inválido o no pertenece a la institución: {programa_id}"
                resultados.append(obj)
                continue

        # parse ints
        horas = 0
        if reg.get("horas_requeridas"):
            try:
                horas = int(float(reg["horas_requeridas"]))
            except Exception:
                horas = 0

        # crear/actualizar user por email
        user = User.objects.filter(email=email).first()
        creado = False

        try:
            if user is None:
                user = User.objects.create(
                    username=username,
                    email=email,
                    first_name=nombre,
                    last_name=apellido,
                )
                if hasattr(user, "phone"):
                    user.phone = telefono
                user.set_password(DEFAULT_PASSWORD)
                user.save()
                creado = True
                lote.creados += 1
            else:
                # update suave
                user.first_name = nombre or user.first_name
                user.last_name = apellido or user.last_name
                if hasattr(user, "phone") and telefono:
                    user.phone = telefono
                user.save()
                lote.actualizados += 1

            # rol practicante
            if hasattr(user, "add_role"):
                user.add_role("practicante")

            # perfil practicante
            perfil, _ = PerfilPracticante.objects.get_or_create(usuario=user)
            perfil.institucion = lote.institucion
            if programa:
                perfil.programa = programa
            if reg.get("periodo_inicio"):
                perfil.periodo_inicio = reg.get("periodo_inicio") or None
            if reg.get("periodo_fin"):
                perfil.periodo_fin = reg.get("periodo_fin") or None
            perfil.horas_requeridas = horas
            perfil.cargado_por_institucion = True
            perfil.save()

            EstudianteImportado.objects.get_or_create(usuario=user, lote=lote)

            obj["estado"] = "ok"
            obj["msg"] = "Creado" if creado else "Actualizado"
            resultados.append(obj)

        except Exception as e:
            lote.con_error += 1
            obj["estado"] = "error"
            obj["msg"] = str(e)
            resultados.append(obj)

    lote.total_registros = total
    lote.log_errores = resultados

    if lote.con_error > 0:
        lote.estado = "completed_with_errors"
    else:
        lote.estado = "completed"

    lote.save(update_fields=["estado", "log_errores", "total_registros", "creados", "actualizados", "con_error"])


class LoteImportacionListView(generics.ListAPIView):
    """GET /api/imports/ — lotes del usuario institución."""
    serializer_class = LoteImportacionSerializer
    permission_classes = [IsInstitucion]

    def get_queryset(self):
        return LoteImportacion.objects.filter(subido_por=self.request.user)


class LoteImportacionUploadView(APIView):
    """POST /api/imports/upload/"""
    permission_classes = [IsInstitucion]

    def post(self, request):
        archivo = request.FILES.get("file")
        if not archivo:
            return Response({"error": "Archivo requerido"}, status=status.HTTP_400_BAD_REQUEST)

        lote = LoteImportacion.objects.create(
            institucion=request.user.institucion,
            subido_por=request.user,
            archivo=archivo,
        )

        # ✅ PROCESAR DE UNA (sin Celery)
        procesar_lote_excel(lote)

        return Response(LoteImportacionSerializer(lote).data, status=status.HTTP_201_CREATED)


class LoteImportacionStatusView(generics.RetrieveAPIView):
    """GET /api/imports/{id}/status/"""
    serializer_class = LoteImportacionSerializer
    permission_classes = [IsInstitucion]

    def get_queryset(self):
        return LoteImportacion.objects.filter(subido_por=self.request.user)